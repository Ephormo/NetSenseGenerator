from flask import Flask, render_template, request, jsonify,url_for
import jieba
import pandas as pd
import pinyin
from openpyxl import load_workbook
import re
import os
import ollama
from pathlib import Path
from typing import List, Dict
import logging
from fuzzywuzzy import fuzz
import asyncio


app = Flask(__name__)

# 初始化jieba分词
jieba.initialize()

def load_dictionary():
    try:
        wb = load_workbook('data/bible_new.xlsx')
        ws = wb.active
        raw_col = None
        cx_col = None

        for col in ws.iter_cols(max_row=1):
            if 'raw' in str(col[0].value).lower():
                raw_col = col[0].column_letter
            elif 'chouxiang' in str(col[0].value).lower():
                cx_col = col[0].column_letter

        if not raw_col or not cx_col:
            raise ValueError("未找到'raw'或'chouxiang'列")

        bible_light_dict = {}
        bible_deep_dict = {}

        for row in ws.iter_rows(min_row=2):
            raw = str(row[ord(raw_col) - 65].value).strip()
            cx = str(row[ord(cx_col) - 65].value).strip()

            if raw and cx:
                bible_light_dict[raw] = cx
                try:
                    py = pinyin.get(raw, format='strip') if not raw.isdigit() else raw
                    bible_deep_dict[py] = cx
                except:
                    bible_deep_dict[raw] = cx

        print("\n字典加载完成，样本检查：")
        for i, (k, v) in enumerate(bible_light_dict.items()):
            if i >= 3: break
            print(f"'{k}' → '{v}'")

        return bible_light_dict, bible_deep_dict

    except Exception as e:
        print(f"\n⚠️ 加载字典出错: {str(e)}")
        return {"你好": "😊", "开心": "😄"}, {"nihao": "😊", "kaixin": "😄"}

bible_light_dict, bible_deep_dict = load_dictionary()

def text_to_emoji(text):
    result = []
    for word in jieba.cut(text):
        word = word.strip()

        if word in bible_light_dict:
            result.append(bible_light_dict[word])
            continue

        try:
            word_py = pinyin.get(word, format='strip')
            if word_py in bible_deep_dict:
                result.append(bible_deep_dict[word_py])
                continue
        except:
            pass

        for char in word:
            if char in bible_light_dict:
                result.append(bible_light_dict[char])
            else:
                try:
                    char_py = pinyin.get(char, format='strip')
                    result.append(bible_deep_dict.get(char_py, char))
                except:
                    result.append(char)

    return ''.join(result)



# Emoji 映射库（可自定义扩展）
EMOJI_MAPPING = {
    '开心': '😊', '高兴': '😄', '快乐': '😀', '笑': '😂',
    '生气': '😠', '愤怒': '🤬', '讨厌': '😤',
    '悲伤': '😢', '难过': '😭', '哭': '😭',
    '惊讶': '😲', '震惊': '🤯', '意外': '😮',
    '爱': '❤️', '喜欢': '🥰', '心': '💖',
    '疑问': '❓', '问题': '🤔', '为什么': '⁉️',
    '时间': '⏰', '早上': '🌅', '晚上': '🌃',
    '天气': '☀️', '雨': '🌧️', '雪': '❄️',
    '吃': '🍔', '喝': '🍹', '美食': '🍕',
    '动物': '🐶', '猫': '🐱', '狗': '🐕',
    '工作': '💼', '学习': '📚', '钱': '💰',
    '默认': '✨', '家': '🏠', '车': '🚗', '飞机': '✈️',
    '火车': '🚆', '船': '🚢', '自行车': '🚲', '公交车': '🚌',
    '地铁': '🚇', '出租车': '🚕', '摩托车': '🏍️', '火箭': '🚀',
    '电话': '📞', '手机': '📱', '电脑': '💻', '电视': '📺',
    '相机': '📷', '音乐': '🎵', '电影': '🎬', '书': '📖',
    '礼物': '🎁', '生日': '🎂', '圣诞': '🎄', '新年': '🎉',
    '派对': '🎊', '运动': '⚽', '篮球': '🏀', '足球': '⚽',
    '网球': '🎾', '棒球': '⚾', '高尔夫': '⛳', '游泳': '🏊',
    '跑步': '🏃', '健身': '🏋️', '瑜伽': '🧘', '拳击': '🥊',
    '滑雪': '⛷️', '滑冰': '⛸️', '冲浪': '🏄', '骑马': '🏇',
    '钓鱼': '🎣', '登山': '🧗', '露营': '🏕️', '旅行': '🧳',
    '地图': '🗺️', '指南针': '🧭', '酒店': '🏨', '餐厅': '🍽️',
    '咖啡': '☕', '茶': '🍵', '啤酒': '🍺', '葡萄酒': '🍷',
    '鸡尾酒': '🍸', '冰淇淋': '🍦', '蛋糕': '🍰', '巧克力': '🍫',
    '糖果': '🍬', '饼干': '🍪', '面包': '🍞', '披萨': '🍕',
    '汉堡': '🍔', '热狗': '🌭', '薯条': '🍟', '寿司': '🍣',
    '拉面': '🍜', '沙拉': '🥗', '水果': '🍎', '蔬菜': '🥦',
    '肉': '🍖', '鱼': '🐟', '虾': '🦐', '螃蟹': '🦀',
    '龙虾': '🦞', '章鱼': '🐙', '贝壳': '🐚', '花': '🌸',
    '树': '🌳', '草': '🌿', '叶子': '🍃', '太阳': '🌞',
    '月亮': '🌜', '星星': '⭐', '云': '☁️', '雨伞': '☂️',
    '彩虹': '🌈', '火': '🔥', '水': '💧', '冰': '🧊',
    '雪人': '⛄', '风': '🌬️', '闪电': '⚡', '山': '⛰️',
    '海': '🌊', '沙漠': '🏜️', '森林': '🌲', '岛': '🏝️',
    '城市': '🏙️', '建筑': '🏢', '桥': '🌉', '塔': '🗼',
    '雕像': '🗽', '教堂': '⛪', '寺庙': '🛕', '清真寺': '🕌',
    '城堡': '🏰', '学校': '🏫', '医院': '🏥', '银行': '🏦',
    '邮局': '🏤', '警察局': '🚓', '消防局': '🚒', '图书馆': '📚',
    '博物馆': '🏛️', '剧院': '🎭', '电影院': '🎦', '商店': '🏬',
    '超市': '🛒', '市场': '🛍️', '公园': '🏞️', '动物园': '🦁',
    '游乐园': '🎡', '水族馆': '🐠', '植物园': '🌺', '农场': '🚜',
    '工厂': '🏭', '机场': '🛫', '火车站': '🚉', '地铁站': '🚇',
    '公交站': '🚏', '加油站': '⛽', '停车场': '🅿️', '厕所': '🚽',
    '浴室': '🛁', '厨房': '🍳', '卧室': '🛏️', '客厅': '🛋️',
    '办公室': '🏢', '会议室': '🏢', '实验室': '🔬', '教室': '🏫',
    '操场': '🏟️', '体育馆': '🏟️', '游泳池': '🏊', '健身房': '🏋️',
    '诊所': '🏥', '药店': '💊', '美容院': '💅', '理发店': '💇',
    '洗衣店': '🧺', '修理店': '🔧', '车库': '🚗', '仓库': '📦',
    '工地': '🏗️', '码头': '🚢', '港口': '⚓', '灯塔': '🚨',
    '信号灯': '🚦', '路标': '🚧', '交通': '🚥', '安全': '🛡️',
    '警告': '⚠️', '禁止': '🚫', '允许': '✅', '帮助': '🆘',
    '急救': '🚑', '健康': '💉', '药品': '💊', '医生': '👨‍⚕️',
    '护士': '👩‍⚕️', '病人': '🤒', '受伤': '🤕', '康复': '💪',
    '运动员': '🏃', '教练': '🏋️', '裁判': '⚖️', '冠军': '🏆',
    '奖牌': '🥇', '奖杯': '🏆', '比赛': '🏅', '训练': '🏋️',
    '团队': '👥', '合作': '🤝', '竞争': '⚔️', '胜利': '🎉',
    '失败': '😞', '努力': '💪', '坚持': '✊', '梦想': '🌟',
    '希望': '🌈', '信念': '🙏', '勇气': '🦁', '智慧': '🧠',
    '知识': '📚', '学习': '📖', '教育': '🏫', '考试': '📝',
    '成绩': '📊', '证书': '🎓', '毕业': '🎓', '工作': '💼',
    '职业': '👔', '事业': '📈', '成功': '🏆', '失败': '😞',
    '挑战': '💪', '机会': '🌟', '风险': '⚠️', '安全': '🛡️',
    '健康': '💉',
    # 没有匹配到关键词时使用的默认emoji
}

def analyze_sentence(sentence):
    sentence_lower = sentence.lower()
    for keyword, emoji in EMOJI_MAPPING.items():
        if keyword in sentence_lower:
            return emoji
    return EMOJI_MAPPING.get('默认', '')

def add_emoji_to_text(text):
    punctuations = r'([。！？，；])'
    parts = re.split(punctuations, text)
    result = []

    for i, part in enumerate(parts):
        if part in ['。', '！', '？', '，', '；']:
            if i > 0 and parts[i - 1].strip():
                sentence = parts[i - 1]
                emoji = analyze_sentence(sentence)
                result[-1] = result[-1] + emoji
            result.append(part)
        else:
            result.append(part)

    return ''.join(result)




# function3

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class EnhancedMemeAssistant:
    def __init__(self, excel_path: str, model: str = "qwen:7b"):
        excel_path='data/memes_dataset.xlsx'
        model='qwen'
        self.excel_path = Path(excel_path)
        self.model = model
        self.meme_db = self._load_meme_database()
        self.all_keywords = self._extract_all_keywords()

    def _load_meme_database(self) -> List[Dict]:
        """加载Excel中的梗数据库"""
        try:
            df = pd.read_excel(self.excel_path)
            # 检查必要列是否存在
            if not all(col in df.columns for col in ["梗名称", "摘要", "标签"]):
                raise ValueError("Excel必须包含'梗名称','摘要','标签'三列")

            # 预处理数据
            records = df.to_dict("records")
            for item in records:
                # 处理标签列，确保是列表格式
                if isinstance(item["标签"], str):
                    # 处理多种分隔符情况
                    tags = re.sub(r"[\[\]'\"]", "", item["标签"])
                    item["标签"] = [tag.strip() for tag in re.split(r"[,，\s]+", tags) if tag.strip()]
                elif pd.isna(item["标签"]):
                    item["标签"] = []
            return records
        except Exception as e:
            logger.error(f"加载Excel文件失败: {e}")
            raise RuntimeError(f"加载Excel文件失败: {e}")

    def _extract_all_keywords(self) -> List[str]:
        """提取所有关键词(梗名称+标签)用于快速匹配"""
        keywords = set()
        for item in self.meme_db:
            keywords.add(item["梗名称"])
            for tag in item["标签"]:
                keywords.add(tag.lower())
        return list(keywords)

    def _find_best_match(self, query: str) -> Dict:
        """找到与查询最匹配的梗(使用模糊匹配)"""
        best_match = None
        highest_score = 0

        for item in self.meme_db:
            # 计算梗名称匹配度
            name_score = fuzz.token_set_ratio(query.lower(), item["梗名称"])

            # 计算标签匹配度(取最高分标签)
            tag_score = max(
                [fuzz.token_set_ratio(query.lower(), tag.lower()) for tag in item["标签"]] or [0]
            )

            # 综合评分(名称权重更高)
            total_score = name_score * 0.7 + tag_score * 0.3

            if total_score > highest_score:
                highest_score = total_score
                best_match = item

        return best_match, highest_score

    def _find_related_matches(self, query: str, exclude: str, threshold: int = 40) -> List[Dict]:
        """找到相关的梗(排除最佳匹配)"""
        related = []

        for item in self.meme_db:
            if item["梗名称"] == exclude:
                continue

            # 计算梗名称匹配度
            name_score = fuzz.token_set_ratio(query.lower(), item["梗名称"])

            # 计算标签匹配度(取最高分标签)
            tag_score = max(
                [fuzz.token_set_ratio(query.lower(), tag.lower()) for tag in item["标签"]] or [0]
            )

            # 综合评分
            total_score = name_score * 0.5 + tag_score * 0.5

            if total_score >= threshold:
                related.append((item, total_score))

        # 按匹配度排序并返回前3个
        related.sort(key=lambda x: x[1], reverse=True)
        return [item for item, score in related[:3]]

    async def ask(self, query: str) -> str:
        """
        向助手提问

        参数:
            query: 用户输入的问题(中文)

        返回:
            助手的回答(包含最佳匹配和相关推荐)
        """
        # 1. 找到最佳匹配
        best_match, match_score = self._find_best_match(query)

        response = ""

        # 如果有较好的匹配(分数>50)
        if best_match and match_score > 50:
            response += f"🎯 最匹配的结果(匹配度{match_score}%):\n"
            response += f"【{best_match['梗名称']}】\n{best_match['摘要']}\n"
            if best_match["标签"]:
                response += f"🏷️ 相关标签: {'、'.join(best_match['标签'])}\n"
            response += "\n"

            # 2. 查找相关推荐
            related_memes = self._find_related_matches(query, best_match["梗名称"])

            if related_memes:
                response += "🔍 您可能还对以下梗感兴趣:\n\n"
                for meme in related_memes:
                    response += f"▪ {meme['梗名称']}: {meme['摘要'][:60]}...\n"
                    if meme["标签"]:
                        response += f"  标签: {'、'.join(meme['标签'][:3])}\n"
                    response += "\n"
        else:
            # 3. 没有足够好的匹配时调用Ollama
            try:
                prompt = (
                    f"你是一个中文网络热梗知识助手。用户问: {query}\n"
                    "请用简洁明了的中文回答关于网络流行梗的问题。"
                    "如果不知道确切答案，可以给出合理的推测或相关梗的介绍。"
                )

                ai_response = await ollama.chat(
                    model=self.model,
                    messages=[{"role": "user", "content": prompt}]
                )
                response = ai_response['message']['content']

                # 即使调用AI也尝试提供一些可能相关的梗
                related_memes = self._find_related_matches(query, "", threshold=30)
                if related_memes:
                    response += "\n\n🔍 以下可能相关的网络梗:\n"
                    for meme in related_memes:
                        response += f"▪ {meme['梗名称']}: {meme['摘要'][:50]}...\n"
            except Exception as e:
                logger.error(f"调用Ollama API失败: {e}")
                response = "抱歉，我暂时无法回答这个问题。您可以尝试换种方式提问。"

                # 仍然尝试提供一些可能相关的梗
                related_memes = self._find_related_matches(query, "", threshold=20)
                if related_memes:
                    response += "\n\n以下可能相关的网络梗:\n"
                    for meme in related_memes:
                        response += f"▪ {meme['梗名称']}\n"

        return response






# flask

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/page1')
def page1():
    return render_template('page1.html')

@app.route('/page2')
def page2():
    return render_template('page2.html')

@app.route('/page3')
def page3():
    return render_template('page3.html')

@app.route('/convert', methods=['POST'])
def convert():
    try:
        data = request.get_json()
        text = data.get('text', '')
        if not text:
            return jsonify({'error': '请输入文本'}), 400

        result = text_to_emoji(text)
        print(f"转换结果: {text} → {result}")
        return jsonify({'result': result})

    except Exception as e:
        print(f"转换出错: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/add_emoji', methods=['POST'])
def add_emoji():
    try:
        data = request.get_json()
        text = data.get('text', '')
        if not text:
            return jsonify({'error': '请输入文本'}), 400

        result = add_emoji_to_text(text)
        print(f"添加Emoji结果: {text} → {result}")
        return jsonify({'result': result})

    except Exception as e:
        print(f"添加Emoji出错: {str(e)}")
        return jsonify({'error': str(e)}), 500


# function3
meme_searcher = EnhancedMemeAssistant("memes_dataset.xlsx")

@app.route('/lookup', methods=['POST'])
def lookup():
    try:
        data = request.get_json()
        query = data.get('text', '')
        if not query:
            return jsonify({'error': '请输入文本'}), 400

        # 使用 asyncio 调用 async 方法
        result = asyncio.run(meme_searcher.ask(query))
        return jsonify({'result': result})

    except Exception as e:
        logger.error(f"查询出错: {str(e)}")
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    os.makedirs('data', exist_ok=True)
    app.run(debug=True, port=5001)





